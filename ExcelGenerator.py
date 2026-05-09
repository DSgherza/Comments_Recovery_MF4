import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime
from DataEntryTemplate import DataEntryTemplate as DET

class ExcelGenerator():
    def __init__(self, _subjective, _objective, link, _subjective_indexes:list, _saving_folder:str, _control:list, _file_name:list, _manouver, check_list_slip:list, coerence:list, report: list, cb, speed_check, target_vs_real_speed_check, direction_check, yaw_check, ay_check, lws_check, excel_bigol_list, signal_lists) -> None:

        self.subjective = _subjective
        self.subjective_indexes = _subjective_indexes
        self.objective = _objective
        
        self.main_repository = _saving_folder
        self.control = _control
        self.file_name = _file_name
        self.maneouvres = [item[0] for item in _manouver]
        self.coerence = coerence

        self.get_database()
        self.get_excel(check_list_slip, report, link, cb, speed_check, target_vs_real_speed_check, direction_check, yaw_check, ay_check, lws_check, signal_lists)
        self.get_bigol_excel(excel_bigol_list)
    
    def get_bigol_excel(self, excel_bigol_list):
        sheet_name_list = []
        df_list = []
        workbook_name = []
        for item in self.file_name:
            name = item.split("/")[-1]
            name = name.replace("mf4", "xlsx")
            name = name.replace("MF4", "xlsx")
            workbook_name.append(name)
        cont = 0
        for item in excel_bigol_list:
            if item[0] == []:
                sheet_name_list.append(item[0])
                df_list.append(item[0])
            else:
                sheet_name_list.append([])
                df_list.append([])
                for ogg in item:
                    sheet_name_list[cont].append(ogg[0])
                    ogg.remove(ogg[0])
                    df_list[cont].append(pd.DataFrame(ogg).T)
            cont += 1
        real_sheet_name_list = []
        cont = 0
        for item in sheet_name_list:
            names_dict = {}
            if item == []:
                real_sheet_name_list.append([])
            else:
                real_sheet_name_list.append([])
                for ogg in item:
                    if ogg not in names_dict:
                        names_dict[ogg] = 1
                    else:
                        names_dict[ogg] = names_dict[ogg] + 1
                    real_sheet_name_list[cont].append(ogg + "_" + str(names_dict[ogg]))
            cont += 1
        for item in df_list:
            if item != []:
                for ogg in item:
                    try:
                        ogg.columns = ['Time', 'BrakePedalPosition', 'Velocity_kmh', 'Ax', 'Ay', 'Ay_calc', 'YawRate', 'YawRate_calc', 'RollRate']
                    except ValueError:
                        ogg.columns = ['Time', 'BrakePedalPosition', 'Velocity_kmh', 'Ax', 'Ay', 'Ay_calc', 'YawRate', 'YawRate_calc']
        cont = 0
        for item in df_list:
            if item != []:
                excel_file = openpyxl.Workbook()
                excel_file.save(self.main_repository + workbook_name[cont])
                excel_file.close()
                contat = 0
                for ogg in item:
                    with pd.ExcelWriter(self.main_repository + workbook_name[cont], engine='openpyxl', mode='a') as writer:
                        ogg.to_excel(writer, sheet_name = real_sheet_name_list[cont][contat], index = False)
                    excel_file = openpyxl.load_workbook(self.main_repository + workbook_name[cont])
                    sheet = excel_file[real_sheet_name_list[cont][contat]]
                    sheet.insert_rows(2)
                    sheet["A2"].value = "[s]"
                    sheet["A2"].alignment = Alignment(horizontal = "center")
                    sheet["B2"].value = "[mm]"
                    sheet["B2"].alignment = Alignment(horizontal = "center")
                    sheet["C2"].value = "[km/h]"
                    sheet["C2"].alignment = Alignment(horizontal = "center")
                    sheet["D2"].value = "[m/s^2]"
                    sheet["D2"].alignment = Alignment(horizontal = "center")
                    sheet["E2"].value = "[m/s^2]"
                    sheet["E2"].alignment = Alignment(horizontal = "center")
                    sheet["F2"].value = "[m/s^2]"
                    sheet["F2"].alignment = Alignment(horizontal = "center")
                    sheet["G2"].value = "[°/s]"
                    sheet["G2"].alignment = Alignment(horizontal = "center")
                    sheet["H2"].value = "[°/s]"
                    sheet["H2"].alignment = Alignment(horizontal = "center")
                    sheet["I2"].value = "[°/s]"
                    sheet["I2"].alignment = Alignment(horizontal = "center")
                    try:    
                        std = excel_file.get_sheet_by_name('Sheet')
                        excel_file.remove_sheet(std)
                    except KeyError:
                        pass
                    excel_file.save(self.main_repository + workbook_name[cont])
                    contat += 1
            cont += 1

    #creare  il database
    def get_database(self):
        self.complete_database =  pd.concat([self.subjective, self.objective.T], axis=1)
        self.complete_database.reset_index(names = "Maneuver", inplace = True)
        indexes_list = []
        for i in range(len(self.subjective_indexes)):
            for j in range(len(self.subjective_indexes[i])):
                indexes_list.append(self.subjective_indexes[i][j])
        self.complete_database.set_index([indexes_list], inplace=True)

    #formattare il file excel
    def get_excel(self, check_list_slip, report, link, cb, speed_check, target_vs_real_speed_check, direction_check, yaw_check, ay_check, lws_check, signal_lists):

        date = str(datetime.now())
        date = date.replace("-", "")
        date = date.replace(":", "")
        date = date.replace(" ", "_")
        date = date.split(".")[0]
        if "/" in self.main_repository:
            output_file_name = self.main_repository + self.main_repository.split("/")[-2] + "_" + date + ".xlsx"
        elif "\\" in self.main_repository:
             output_file_name = self.main_repository + self.main_repository.split("\\")[-2] + "_" + date + ".xlsx"
        output_file_name = self.main_repository + self.main_repository.split("/")[-2] + "_" + date + ".xlsx"
        DET(self.complete_database, cb, link, output_file_name)
        self.complete_database.to_excel(output_file_name, sheet_name = "Global Analysis")
        file_excel = openpyxl.load_workbook(output_file_name)
        signal_excel = openpyxl.load_workbook(os.getcwd() + r'\signalList.xlsx')
        for item in signal_lists:
            signal_sheet = signal_excel[item]
            file_excel.create_sheet(title = item)
            file_sheet = file_excel[item]
            r = 0
            for row in signal_sheet.rows:
                c = 0
                r = r + 1
                for cell in row:
                    c = c + 1
                    file_sheet.cell(r, c, value = cell.value)
                    if signal_sheet.cell(r, c).fill.fgColor.value == 'FF00B050':
                        file_sheet.cell(r, c).fill = PatternFill(start_color = 'FF00B050', end_color = 'FF00B050', fill_type = "solid")
        signal_excel.close()
        sheet = file_excel["Global Analysis"]
        for col in sheet.columns:
            for cell in col:
                cell.font = Font(name = "Allumi Pro", size = 10)
        for col in ["DU", "DV", "DW", "DX", "DY", "DZ"]:
            counter = 0
            column = sheet[col]
            for cell in column:
                if cell.row != 1:
                    try:
                        if check_list_slip[counter]:
                           cell.fill = PatternFill(start_color = "D2B48C", end_color = "D2B48C", fill_type = "solid")
                    except IndexError:
                        break
                    counter += 1
        counter = 0
        column = sheet["B"]
        for cell in column:
            if cell.row != 1:
                try:
                    if self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    pass
                try:
                    if speed_check[counter] == "Internal":
                        cell.fill = PatternFill(start_color = "FF00FF", end_color = "FF00FF", fill_type = "solid")
                    elif speed_check[counter] == "TestaOttica":
                        cell.fill = PatternFill(start_color = "99CBFF", end_color = "99CBFF", fill_type = "solid")
                except IndexError:
                    pass
                counter += 1
        counter = 0
        column = sheet["N"]
        for cell in column:
            if cell.row != 1:
                try:
                    if self.coerence[counter] == 1:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        counter = 0
        column = sheet["O"]
        for cell in column:
            if cell.row != 1:
                try:
                    if self.coerence[counter] == 1:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        counter = 0
        column = sheet["T"]
        for cell in column:
            if cell.row != 1:
                try:
                    if target_vs_real_speed_check[counter] == False:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        counter = 0
        column = sheet["AX"]
        for cell in column:
            if cell.row != 1:
                try:
                    if ay_check[counter] == False:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        counter = 0
        column = sheet["BN"]
        for cell in column:
            if cell.row != 1:
                try:
                    if yaw_check[counter] == False:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        counter = 0
        column = sheet["CF"]
        for cell in column:
            if cell.row != 1:
                try:
                    if lws_check[counter] == False:
                       cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                except IndexError:
                    break
                counter += 1
        columns_values = [[["ABS_Chessboard", "ABS_NegativeMuSplitJump", "ABS_PositiveMuSplitJump", "ABS_MuSplit"], [77, 85, 88], [7, 90, 100]], 
                          [["ABS_StoppingDistance", "ABS_NegativeMuJump", "ABS_PositiveMuJump", "ABS_DoubleMuJump", "TCS_LaunchHomogeneous", "TCS_PositiveMuJump", "TCS_NegativeMuJump"], [77], [4]],
                          [["ABS_LaneChange", "ESC_RampSteer", "ESC_LaneChange", "ESC_Slalom", "ESC_ConstantRadius", "ESC_Handling"], [91], [10]],
                          [["ESC_StepSteer"], [78, 91], [8, 10]], 
                          [["TCS_Chessboard", "TCS_NegativeMuSplitJump", "TCS_PositiveMuSplitJump", "TCS_LaunchMuSplit"], [77, 85, 88], [4, 45, 50]],
                          [["ABS_InTurn", "ESC_PartialBrkinTurn", "ESC_PowerOffinTurn", "ESC_PowerOninTurn"], [77, 91], [4, 10]]]
        column = sheet["B"]
        red_values = []
        for cell in column:
            if cell.value in ["ABS_Chessboard", "ABS_InTurn", "ABS_LaneChange", "ABS_MuSplit", "ABS_NegativeMuJump", "ABS_NegativeMuSplitJump", 
                                              "ABS_PositiveMuJump", "ABS_PositiveMuSplitJump", "ABS_StoppingDistance", "ABS_DoubleMuJump", "ESC_ConstantRadius", 
                                              "ESC_Handling", "ESC_LaneChange", "ESC_PartialBrkinTurn", "ESC_PowerOffinTurn", "ESC_PowerOninTurn", "ESC_RampSteer", 
                                              "ESC_Slalom", "ESC_StepSteer", "TCS_Chessboard", "TCS_LaunchHomogeneous", "TCS_LaunchMuSplit", "TCS_NegativeMuJump", 
                                              "TCS_NegativeMuSplitJump", "TCS_PositiveMuJump", "TCS_PositiveMuSplitJump"]:
                red_values_list = []
                for i in range(len(columns_values)):
                    if cell.value in columns_values[i][0]:
                        for j in range(len(columns_values[i][1])):
                            aux_cell = sheet.cell(row = cell.row, column = columns_values[i][1][j])
                            try:
                                if abs(aux_cell.value) > columns_values[i][2][j]:
                                    aux_cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                                    red_values_list.append(False)
                                else:
                                    red_values_list.append(True)
                            except TypeError:
                                red_values_list.append(True)
                                pass
                        found = False
                        for j in range(len(red_values_list)):
                            if not red_values_list[j]:
                                red_values.append(False)
                                found = True
                                break
                        if not found:
                            red_values.append(True)
            elif cell.value == "ABS_Braking":
                red_values.append(True)
        counter = 0
        for cell in sheet["B"]:
            if cell.value in ["ABS_Chessboard", "ABS_InTurn", "ABS_LaneChange", "ABS_MuSplit", "ABS_NegativeMuJump", "ABS_NegativeMuSplitJump", 
                                              "ABS_PositiveMuJump", "ABS_PositiveMuSplitJump", "ABS_StoppingDistance", "ABS_DoubleMuJump", "ESC_ConstantRadius", 
                                              "ESC_Handling", "ESC_LaneChange", "ESC_PartialBrkinTurn", "ESC_PowerOffinTurn", "ESC_PowerOninTurn", "ESC_RampSteer", 
                                              "ESC_Slalom", "ESC_StepSteer", "TCS_Chessboard", "TCS_LaunchHomogeneous", "TCS_LaunchMuSplit", "TCS_NegativeMuJump", 
                                              "TCS_NegativeMuSplitJump", "TCS_PositiveMuJump", "TCS_PositiveMuSplitJump"]:
                if (self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False) and red_values[counter] == False and speed_check[counter] == "VBOX":
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                elif red_values[counter] == False and (self.coerence[counter] == 2 and (target_vs_real_speed_check[counter] == True or target_vs_real_speed_check[counter] == "") and (direction_check[counter] == True or direction_check[counter] == "")) and speed_check[counter] == "VBOX":
                    cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                elif (self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False) and red_values[counter] == False and speed_check[counter] == "TestaOttica":
                    cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 3)
                    aux_cell.fill = PatternFill(start_color = "99CBFF", end_color = "99CBFF", fill_type = "solid")
                elif (self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False) and red_values[counter] == False and speed_check[counter] == "Internal":
                    cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 3)
                    aux_cell.fill = PatternFill(start_color = "FF00FF", end_color = "FF00FF", fill_type = "solid")
                elif (self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False) and speed_check[counter] == "Internal" and red_values[counter] == True:
                    cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "FF00FF", end_color = "FF00FF", fill_type = "solid")
                elif (self.coerence[counter] == 1 or target_vs_real_speed_check[counter] == False or direction_check[counter] == False) and speed_check[counter] == "TestaOttica" and red_values[counter] == True:
                    cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "99CBFF", end_color = "99CBFF", fill_type = "solid")
                elif (self.coerence[counter] == 2 or (target_vs_real_speed_check[counter] == True or target_vs_real_speed_check[counter] == "") or (direction_check[counter] == True or direction_check[counter] == "")) and (speed_check[counter] == "Internal" or speed_check[counter] == "TestaOttica") and red_values[counter] == False:
                    aux_cell = sheet.cell(row = cell.row, column = 1)
                    aux_cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                counter += 1
            elif cell.value == "ABS_Braking":
                counter += 1
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            column = sheet[col]
            for cell in column:
                cell.border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
                if col != "B":
                    cell.alignment = Alignment(horizontal="center")
                    if cell.coordinate in ["C1", "D1", "E1", "F1"]:
                        cell.font = Font(bold = True, size = 11)
                else:
                    if cell.coordinate == "B1":
                        cell.font = Font(bold = True, size = 11)
                if col not in ["A", "B", "C", "D", "E", "F"]:
                    if cell.value == "NR":
                        cell.font = Font(color = "0066CC", bold = True, name = "Allumi Pro", size = 10)
                    if cell.value == 1:
                        cell.fill = PatternFill(start_color = "E53138", end_color="E53138", fill_type="solid")
                    if cell.value == 2:
                        cell.fill = PatternFill(start_color = "E93F57", end_color="E93F57", fill_type="solid")
                    if cell.value == 3:
                        cell.fill = PatternFill(start_color = "F9AE38", end_color="F9AE38", fill_type="solid")
                    if cell.value == 4 or cell.value == 5:
                        cell.fill = PatternFill(start_color = "FAE702", end_color="FAE702", fill_type="solid")
                    if cell.value == 6:
                        cell.fill = PatternFill(start_color = "E7F994", end_color="E7F994", fill_type="solid")
                    if cell.value == 7 or cell.value == 8 or cell.value == 9:
                        cell.fill = PatternFill(start_color = "A2F377", end_color="A2F377", fill_type="solid")
                    if cell.value == 10:
                        cell.fill = PatternFill(start_color = "51FD01", end_color="51FD01", fill_type="solid")
        cont = 0
        sheet.insert_rows(2)
        sheet.merge_cells("A2:C2")
        sheet["A2"].border = Border(left = Side(style = "thin"))
        sheet["F2"].border = Border(right = Side(style = "thin"))
        sheet.insert_rows(3)
        sheet.merge_cells("A3:C3")
        if self.control[0]:
            sheet["A3"].value = '=HYPERLINK("{}", "{}")'.format(link, self.file_name[0].split("/")[-1]) 
            sheet["A3"].font = Font(name = "Allumi Pro", size = 10)
            file_cont = 1
        sheet["A3"].alignment = Alignment(horizontal = "center")
        for item in ["B3", "C3", "D3", "E3"]:
            sheet[item].border = Border(top = Side(style = "thin"))
        sheet["A3"].border = Border(top = Side(style = "thin"), left = Side(style = "thin"))
        sheet["F3"].border = Border(top = Side(style = "thin"), right = Side(style = "thin"))
        cell_to_exclude = [sheet["A3"]]
        n_rows = 0
        for cell in sheet["A"]:
            cont += 1
            if cell.value == 1 and cell.coordinate != "A4":
                cont += 2
                sheet.insert_rows(cont-2)
                sheet.insert_rows(cont-2)
                sheet.merge_cells(start_row=cont-1, start_column=1, end_row=cont-1, end_column=3)
                sheet.merge_cells(start_row=cont-2, start_column=1, end_row=cont-2, end_column=3)
                
                # altri file della lista
                name_file_cell = sheet.cell(row=cont-1, column=1)
                cell_to_exclude.append(name_file_cell)
                if self.control[file_cont]:
                    name_file_cell.value = '=HYPERLINK("{}", "{}")'.format(link, self.file_name[file_cont].split("/")[-1])
                    name_file_cell.font = Font(name = "Allumi Pro", size = 10)
                name_file_cell.alignment = Alignment(horizontal = "center")
                name_file_cell.border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
                aux_cell = sheet.cell(row=cont-1, column=6)
                aux_cell.border = Border(right = Side(style = "thin"))
                for i in range(1,7):
                    if i != 1 and i != 6:
                        aux_cell = sheet.cell(row=cont-2, column=i)
                        aux_cell.border = Border(bottom = Side(style = "thin"))
                    else:
                        if i == 1:
                            aux_cell = sheet.cell(row=cont-2, column=i)
                            aux_cell.border = Border(left = Side(style = "thin"))
                        if i == 6:
                            aux_cell = sheet.cell(row=cont-2, column=i)
                            aux_cell.border = Border(right = Side(style = "thin"), bottom = Side(style = "thin"))
                file_cont += 1
        for cell in sheet["A"]:
            n_rows += 1
        for col in sheet.columns:
            set_len = 0
            column = col[0].column_letter
            for cell in col:
                if cell not in cell_to_exclude:
                    if len(str(cell.value)) > set_len:
                        set_len = len(str(cell.value))
                        set_col_width = set_len + 3
                sheet.column_dimensions[column].width = set_col_width
        color_list = ["E53138", "E93F57", "F9AE38", "FAE702", "FAE702", "E7F994", "A2F377", "A2F377", "A2F377", "51FD01"]
        cell_list = [sheet.cell(n_rows + 4, 5), sheet.cell(n_rows + 4, 6), sheet.cell(n_rows + 4, 7), sheet.cell(n_rows + 4, 8), sheet.cell(n_rows + 4, 9), sheet.cell(n_rows + 4, 10), sheet.cell(n_rows + 4, 11), sheet.cell(n_rows + 4, 12), sheet.cell(n_rows + 4, 13), sheet.cell(n_rows + 4, 14)]
        sheet.merge_cells(start_row = n_rows + 3, start_column = 5, end_row = n_rows + 3, end_column = 14)
        sheet.cell(n_rows + 3, 5).value = "SAE J1441-201609 COLOR SCALE"
        sheet.cell(n_rows + 3, 5).alignment = Alignment(horizontal = "center")
        sheet.cell(n_rows + 3, 5).font = Font(name = "Allumi Pro", size = 10)
        for i in range(len(cell_list)):
            cell_list[i].value = i+1
            cell_list[i].fill = PatternFill(start_color = color_list[i], end_color=color_list[i], fill_type="solid")
            cell_list[i].border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            cell_list[i].alignment = Alignment(horizontal = "center")
            cell_list[i].font = Font(name = "Allumi Pro", size = 10)
        sheet.merge_cells(start_row = n_rows + 7, start_column = 8, end_row = n_rows + 7, end_column = 11)    
        sheet.cell(n_rows + 7, 8).value = "NR = Not Registered"
        sheet.cell(n_rows + 7, 8).font = Font(color = "0066CC", bold = True, name = "Allumi Pro", size = 10)
        sheet.cell(n_rows + 7, 8).border = Border(left = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        sheet.cell(n_rows + 7, 11).border = Border(right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        sheet.cell(n_rows + 7, 9).border = Border(top = Side(style = "thin"), bottom = Side(style = "thin"))
        sheet.cell(n_rows + 7, 10).border = Border(top = Side(style = "thin"), bottom = Side(style = "thin"))
        sheet.cell(n_rows + 7, 8).alignment = Alignment(horizontal = "center")
        sheet.cell(n_rows + 9, 2).value = "Colors legend:"
        sheet.cell(n_rows + 9, 2).font = Font(name = "Allumi Pro", size = 10)
        sheet.cell(n_rows + 10, 2).fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
        sheet.cell(n_rows + 10, 3).value = "Calculated value over expected range"
        sheet.cell(n_rows + 10, 3).font = Font(name = "Allumi Pro", size = 10)
        sheet.cell(n_rows + 11, 2).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
        sheet.cell(n_rows + 11, 3).value = "Maneuver not correctly performed"
        sheet.cell(n_rows + 11, 3).font = Font(name = "Allumi Pro", size = 10)
        sheet.cell(n_rows + 12, 2).fill = PatternFill(start_color = "99CBFF", end_color = "99CBFF", fill_type = "solid")
        sheet.cell(n_rows + 12, 3).value = "SMotion Speed used in calculations"
        sheet.cell(n_rows + 12, 3).font = Font(name = "Allumi Pro", size = 10)
        sheet.cell(n_rows + 13, 2).fill = PatternFill(start_color = "FF00FF", end_color = "FF00FF", fill_type = "solid")
        sheet.cell(n_rows + 13, 3).value = "Internal Speed used in calculations"
        sheet.cell(n_rows + 13, 3).font = Font(name = "Allumi Pro", size = 10)
        for col in sheet.columns:
            if col[0].column_letter not in ["A", "B", "C", "D", "E", "F"]:
                for i in range(len(col)):
                    if i < len(col) - 13:
                        col[i].border = Border(bottom = Side(style = "thin"), right = Side(style = "thin"))
                        col[i].alignment = Alignment(horizontal = "center")
        i = 0
        for cell in sheet["B"]:
            if cell.value in ["ABS_Chessboard", "ABS_InTurn", "ABS_LaneChange", "ABS_MuSplit", "ABS_NegativeMuJump", "ABS_NegativeMuSplitJump", 
                              "ABS_PositiveMuJump", "ABS_PositiveMuSplitJump", "ABS_StoppingDistance", "ABS_DoubleMuJump", "ESC_ConstantRadius", "ESC_Handling",
                              "ESC_LaneChange", "ESC_PartialBrkinTurn", "ESC_PowerOffinTurn", "ESC_PowerOninTurn", "ESC_RampSteer", "ESC_Slalom", 
                              "ESC_StepSteer", "TCS_Chessboard", "TCS_LaunchHomogeneous", "TCS_LaunchMuSplit", "TCS_NegativeMuJump", 
                              "TCS_NegativeMuSplitJump", "TCS_PositiveMuJump", "TCS_PositiveMuSplitJump", "ZERO_Static", "ZERO_Dynamic", "ABS_Braking"]:
                cell.value = '=HYPERLINK("{}", "{}")'.format(report[i], cell.value)
                i += 1
        sheet.freeze_panes = 'D2'
        for item in signal_lists:
            list_sheet = file_excel[item]    
            for col in list_sheet.columns:
                set_len = 0
                column = col[0].column_letter
                for cell in col:
                    if len(str(cell.value)) > set_len:
                        set_len = len(str(cell.value))
                        set_col_width = set_len + 3
                list_sheet.column_dimensions[column].width = set_col_width
        for row in sheet.rows:
            for cell in row:
                cell.font = Font(name = "Allumi Pro", size = 10, bold = True)
            break
        file_excel.save(output_file_name)
        if cb.get() == 1:
            file_excel.save(link + "\\" + output_file_name.split("/")[-1])
        
