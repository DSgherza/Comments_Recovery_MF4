import asammdf
import pandas as pd
import openpyxl
import numpy as np
import scipy

class InputManagement:

    def get_to_filter(self, workbook, vehicle):

        signals_to_filter = []
        wb = openpyxl.load_workbook(workbook, rich_text= True)
        worksheet = wb[vehicle]

        #we check just the column "D" because of the spreadsheet rule of construction
        for row in range(2,worksheet.max_row+1):  
            for column in "D":  
                cell_name = "{}{}".format(column, row)
                if worksheet[cell_name].fill.fgColor.value == 'FF00B050':
                    signals_to_filter.append(worksheet[cell_name].value)

        return signals_to_filter

    def get_priority(self, workbook, vehicle) -> list:
        priority1 = list()
        priority2 = list()
        priority3 = list()
        priority4 = list()
        priority5 = list()

        wb = openpyxl.load_workbook(workbook, rich_text= True)
        worksheet = wb[vehicle]

        for row in range(2,worksheet.max_row+1):  
            for column in "D":  
                cell_name = "{}{}".format(column, row)
                a = worksheet[cell_name]
                # Prio 1
                if worksheet[cell_name].font.sz == 20:
                    priority1.append(worksheet[cell_name].value)
                # Prio 2
                elif worksheet[cell_name].font.sz == 18:
                    priority2.append(worksheet[cell_name].value)
                # Prio 3
                elif worksheet[cell_name].font.sz == 16:
                    priority3.append(worksheet[cell_name].value)
                # Prio 4
                elif worksheet[cell_name].font.sz == 14:
                    priority4.append(worksheet[cell_name].value)        
                # Prio 5
                elif worksheet[cell_name].font.sz == 12:
                    priority5.append(worksheet[cell_name].value)   
                
        return priority1, priority2, priority3, priority4, priority5

    def mdf_variable_names(self, file: asammdf.MDF, sources: list = None, signals: list = None, local_signals: list = None) -> dict:

        variable_names = {}
        times = dict()
        constant_values = ["PressureForceRatio_FA", "PressureForceRatio_RA", "Yield", "Coasting", "RollingRadius_FA", "RollingRadius_RA", "EffectiveRadius_FA", "EffectiveRadius_RA", "Weight"]

        # compare the signal name we want and the signal name in file. if the signal considered is a time channel, it will be saved in 
        # times dictionary. The key is the name of signal.
        #  If the signal is not a time channel, it will be saved in variable_names. The key is the variable name. 

        for i in range(len(local_signals)):
            if local_signals[i] in constant_values:
                variable_names[local_signals[i]] = float(signals[i])
            else:
                try:
                    if len(file.channels_db[local_signals[i]]) == 1:
                        signal_examinated = file.get(local_signals[i])
                        if len(signal_examinated.samples) != 0:
                            temp_dict = {}
                            temp_dict["time"] = signal_examinated.timestamps
                            temp_dict["data"] = signal_examinated.samples
                            times[signals[i]] = signal_examinated.timestamps
                            variable_names[signals[i]] = temp_dict
                    elif len(file.channels_db[local_signals[i]]) == 0:
                        pass
                    else:
                        for j in range(len(file.channels_db[local_signals[i]])):
                            found = False
                            if file.groups[file.channels_db[local_signals[i]][j][0]].channel_group.acq_name == sources[i]:
                                group = file.channels_db[local_signals[i]][j][0]
                                index = file.channels_db[local_signals[i]][j][1]
                                found = True
                                break
                            if found:
                                break
                            if not found:
                                if file.groups[file.channels_db[local_signals[i]][j][0]].channel_group.acq_source.path == sources[i]:
                                    group = file.channels_db[local_signals[i]][j][0]
                                    index = file.channels_db[local_signals[i]][j][1]
                                    break
                        signal_examinated = file.get(local_signals[i], group, index)
                    if len(signal_examinated.samples) != 0:
                        temp_dict = {}
                        temp_dict["time"] = signal_examinated.timestamps
                        temp_dict["data"] = signal_examinated.samples
                        times[signals[i]] = signal_examinated.timestamps
                        variable_names[signals[i]] = temp_dict
                except KeyError:
                    pass
                except AttributeError:
                    occurrences = file.whereis(local_signals[i])
                    group_index, channel_index = occurrences[0]
                    signal_examinated = file.get(local_signals[i], group_index, channel_index)
                    if len(signal_examinated.samples) != 0:
                        temp_dict = {}
                        temp_dict["time"] = signal_examinated.timestamps
                        temp_dict["data"] = signal_examinated.samples
                        times[signals[i]] = signal_examinated.timestamps
                        variable_names[signals[i]] = temp_dict
    
        return variable_names, times

    def get_subjective_signals(self, sources: list = None, signals: list = None,) -> list:

        subjective_signals = list()
        for i in range(len(signals)):
            if sources[i] == "Environment" or sources[i] == "Polling":
                subjective_signals.append(signals[i])
    
        return subjective_signals


    def get_vehicle_recognition (self, vehicle: str, vehicle_list) -> str:
        vehicle_name = vehicle_list["Vehicle Name"].tolist()
        sheet_name = vehicle_list["Sheet Name"].tolist()
        for i in range(len(vehicle_name)):
            if vehicle_name[i] == vehicle:
                vehicle = sheet_name[i]
                cont = i
                break
        cont = cont + 1
        return vehicle, cont

    def __init__(self, filename):
        mdf_file = asammdf.MDF(filename)
        self.log = filename.split("/")[-1]
        self.comment = mdf_file.header.description.replace("\n",". ")

    def diadem_names(self, mdf_input_reference, vehicle_name, lola, times):
        diadem = {}
        mdf_input_name = pd.read_excel(mdf_input_reference, header = 0, sheet_name = vehicle_name)
        mdf_diadem = mdf_input_name["DiademName"]
        mdf_lola = mdf_input_name["VariableName"]
        mdf_units = mdf_input_name["Unit"]
        for i in range(len(mdf_diadem)):
            try:
                if isinstance(mdf_diadem[i], str):
                    diadem[mdf_diadem[i]] = (lola[mdf_lola[i]], mdf_units[i])
            except KeyError:
                pass
        data = [times]
        columns = ["Time"]
        units = ["s"]
        for key in diadem:
            columns.append(key)
            units.append(diadem[key][1])
            data.append(diadem[key][0].tolist())
        diadem_df = pd.DataFrame(data).T
        units_df = pd.DataFrame(units).T
        export_diadem = pd.concat([units_df, diadem_df], axis = 0)
        export_diadem.columns = columns
        export_diadem.reset_index(inplace = True)
        export_diadem.iloc[0,0] = ""
        export_diadem.set_index("index", inplace = True)
        return export_diadem 

    def get_time_channel(self, mdf_times, signals, frequency_user, frequency_user_let, ode, let):

        broken_channels = []
        last_timevalue = []
        first_timevalue = []
        frequency = []
        for key in mdf_times:
            if key != "Comment":
                first_timevalue.append(int(round(mdf_times[key][0])))
                last_timevalue.append(int(round(mdf_times[key][-1])))
        ending_time = max(np.array(last_timevalue))
        for key in mdf_times:
            if key != "Comment":
                if ending_time <= int(round(mdf_times[key][-1])) + 1 and ending_time >= int(round(mdf_times[key][-1])) - 1:
                    pass
                else:
                    broken_channels.append(key)
        
        min_time = 0
        max_time = float("inf")
        for key in mdf_times:
            if key != "Comment" and key not in broken_channels:
                if mdf_times[key][0] > min_time:
                    min_time = mdf_times[key][0]
                if mdf_times[key][-1] < max_time:
                    max_time = mdf_times[key][-1]

        for key in mdf_times:
            if key != "Comment":
                while mdf_times[key][0] < min_time:
                    mdf_times[key] = np.delete(mdf_times[key], 0)
                    signals[key]["time"] = np.delete(signals[key]["time"], 0)
                    signals[key]["data"] = np.delete(signals[key]["data"], 0)
            if key != "Comment" and key not in broken_channels:
                while mdf_times[key][-1] > max_time:
                    mdf_times[key] = np.delete(mdf_times[key], -1)
                    signals[key]["time"] = np.delete(signals[key]["time"], -1)
                    signals[key]["data"] = np.delete(signals[key]["data"], -1)
                
        for key in mdf_times:
            if key != "Comment":
                frequency.append(int(round(1/((mdf_times[key][-1] - mdf_times[key][0])/len(mdf_times[key])))))
        
        if ode.get() == 1:
            if frequency_user.get() == "MAX":
                time_elements = int(round((max_time - min_time)*max(np.array(frequency))))
                self.time_resampled = np.linspace(start = min_time, stop = max_time, num = time_elements)
                self.frequency = max(np.array(frequency))
            else:
                time_elements = int(round((max_time - min_time)*int(frequency_user.get())))
                self.time_resampled = np.linspace(start = min_time, stop = max_time, num = time_elements)
                self.frequency = int(frequency_user.get())
        elif ode.get() == 0 and let.get() == 0:
            time_elements = int(round((max_time - min_time)*max(np.array(frequency))))
            self.time_resampled = np.linspace(start = min_time, stop = max_time, num = time_elements)
            self.frequency = max(np.array(frequency))
        elif let.get() == 1:
            if frequency_user_let.get() == "MAX":
                time_elements = int(round((max_time - min_time)*max(np.array(frequency))))
                self.time_resampled = np.linspace(start = min_time, stop = max_time, num = time_elements)
                self.frequency = max(np.array(frequency))
            else:
                time_elements = int(round((max_time - min_time)*max(np.array(frequency))))
                self.time_resampled = np.linspace(start = min_time, stop = max_time, num = time_elements)
                self.frequency = max(np.array(frequency))
                time_elements_export = int(round((max_time - min_time)*int(frequency_user_let.get())))
                self.time_resampled_export = np.linspace(start = min_time, stop = max_time, num = time_elements_export)
                self.frequency_export = int(frequency_user_let.get())

        return broken_channels, min_time
        
    def get_resampled_data(self, mdf_signals, subjective_signals, signals_to_filter, broken_channels, min_time, let, frequency_let):
        
        b,a = scipy.signal.butter(2, 10/(0.5*167), analog=False, output='ba')    

        for key in mdf_signals:
            constant_values = ["PressureForceRatio_FA", "PressureForceRatio_RA", "Yield", "Coasting", "RollingRadius_FA", "RollingRadius_RA", "EffectiveRadius_FA", "EffectiveRadius_RA", "Weight"]
            if key != "Comment" and key not in broken_channels:
                if key in constant_values:
                    self.objectives[key] = mdf_signals[key]
                else:            
                    if key in subjective_signals:
                        self.subjectives[key] = np.interp(self.time_resampled, mdf_signals[key]['time'],  mdf_signals[key]['data'])
                    else:
                        temp_objective = np.interp(self.time_resampled, mdf_signals[key]['time'], mdf_signals[key]['data'])
                        if key in signals_to_filter:
                            self.objectives[key] = scipy.signal.filtfilt(b, a, temp_objective)
                        else:
                            self.objectives[key] = temp_objective
            if key in broken_channels:
                time_elements = int(round((mdf_signals[key]['time'][-1] - min_time)*self.frequency))
                time_broken = np.linspace(start = min_time, stop = mdf_signals[key]['time'][-1], num = time_elements)
                self.times[key] = [time_broken, time_elements]
                if key in subjective_signals:
                    self.subjectives[key] = np.interp(time_broken, mdf_signals[key]['time'],  mdf_signals[key]['data'])
                else:
                    temp_objective = np.interp(time_broken, mdf_signals[key]['time'], mdf_signals[key]['data'])
                    if key in signals_to_filter:
                        self.objectives[key] = scipy.signal.filtfilt(b, a, temp_objective)
                    else:
                        self.objectives[key] = temp_objective
        
        if let.get() == 1 and frequency_let.get() != "MAX":
            for key in mdf_signals:
                constant_values = ["PressureForceRatio_FA", "PressureForceRatio_RA", "Yield", "Coasting", "RollingRadius_FA", "RollingRadius_RA", "EffectiveRadius_FA", "EffectiveRadius_RA", "Weight"]
                if key != "Comment" and key not in broken_channels:
                    if key in constant_values:
                        self.objectives_export[key] = mdf_signals[key]
                    else:            
                        if key in subjective_signals:
                            self.subjectives_export[key] = np.interp(self.time_resampled_export, mdf_signals[key]['time'],  mdf_signals[key]['data'])
                        else:
                            temp_objective = np.interp(self.time_resampled_export, mdf_signals[key]['time'], mdf_signals[key]['data'])
                            if key in signals_to_filter:
                                self.objectives_export[key] = scipy.signal.filtfilt(b, a, temp_objective)
                            else:
                                self.objectives_export[key] = temp_objective
                if key in broken_channels:
                    time_elements = int(round((mdf_signals[key]['time'][-1] - min_time)*self.frequency_export))
                    time_broken = np.linspace(start = min_time, stop = mdf_signals[key]['time'][-1], num = time_elements)
                    self.times_export[key] = [time_broken, time_elements]
                    if key in subjective_signals:
                        self.subjectives_export[key] = np.interp(time_broken, mdf_signals[key]['time'],  mdf_signals[key]['data'])
                    else:
                        temp_objective = np.interp(time_broken, mdf_signals[key]['time'], mdf_signals[key]['data'])
                        if key in signals_to_filter:
                            self.objectives_export[key] = scipy.signal.filtfilt(b, a, temp_objective)
                        else:
                            self.objectives_export[key] = temp_objective